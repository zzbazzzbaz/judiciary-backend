"""Cases 子应用 Admin 配置。"""

from __future__ import annotations

import os

from django import forms
from django.conf import settings
from django.contrib import admin, messages
from django.db import IntegrityError, transaction
from django.http import FileResponse
from django.shortcuts import redirect
from django.urls import path
from django.utils import timezone
from django.utils.html import format_html
from import_export.admin import ImportMixin
from import_export.formats.base_formats import XLSX
from import_export.results import RowResult

from apps.common.models import Attachment
from apps.grids.models import Grid
from apps.users.models import User
from config.admin_sites import admin_site, grid_manager_site


def get_attachments_from_ids(ids_str: str) -> list:
    """根据逗号分隔的ID字符串获取附件列表。"""
    if not ids_str:
        return []
    try:
        ids = [int(i.strip()) for i in ids_str.split(",") if i.strip()]
        return list(Attachment.objects.filter(id__in=ids))
    except (ValueError, TypeError):
        return []


from utils.code_generator import generate_task_code

from .models import ArchivedTask, CaseArchive, Task, TaskStatReport, TaskType, Town, UnassignedTask
from .resources import CaseArchiveResource


class TaskAdminForm(forms.ModelForm):
    """任务管理表单（管理员）。"""

    class Meta:
        model = Task
        fields = "__all__"

    def clean(self):
        cleaned_data = super().clean()

        reporter: User | None = cleaned_data.get("reporter")
        if reporter and reporter.role != User.Role.MEDIATOR:
            raise forms.ValidationError("上报人必须为调解员")

        grid: Grid | None = cleaned_data.get("grid")
        assigned_mediator: User | None = cleaned_data.get("assigned_mediator")
        if assigned_mediator:
            if assigned_mediator.role != User.Role.MEDIATOR or not assigned_mediator.is_active:
                raise forms.ValidationError("被分配人员必须为启用状态的调解员")
            if not grid:
                raise forms.ValidationError("分派调解员前必须先选择所属网格")
            if assigned_mediator.grid_id != grid.id:
                raise forms.ValidationError("该调解员不在此网格内，请先将调解员分配到该网格")

        return cleaned_data


class TaskAdmin(admin.ModelAdmin):
    """任务管理（纠纷/法律援助、分派与调解结果）。"""

    form = TaskAdminForm
    actions = ["archive_tasks"]
    autocomplete_fields = ("reporter", "assigner", "assigned_mediator")

    class Media:
        js = ("admin/js/task_grid_filter.js",)

    list_display = (
        "id",
        "code",
        "task_type",
        "status",
        "party_name",
        "grid",
        "reporter",
        "assigned_mediator",
        "reported_at",
        "view_detail_action",
    )
    list_select_related = ("grid", "reporter", "assigned_mediator", "task_type")
    search_fields = ("code", "party_name", "party_phone", "reporter__name", "assigned_mediator__name")
    list_filter = ("task_type", "status", "grid", "reported_at")
    ordering = ("-reported_at", "-id")
    list_per_page = 20
    show_full_result_count = True

    readonly_fields = ("code", "status", "reported_at", "created_at", "updated_at")
    fieldsets = (
        ("任务信息", {"fields": ("code", "task_type", "town", "status", "grid", "description", "amount")}),
        ("当事人信息", {"fields": ("party_name", "party_phone", "party_address")}),
        (
            "上报信息",
            {
                "fields": (
                    "reporter",
                    "reported_at",
                    ("report_lng", "report_lat"),
                    "report_address",
                    "report_image_ids",
                    "report_file_ids",
                )
            },
        ),
        ("分派信息", {"fields": ("assigner", "assigned_mediator", "assigned_at")}),
        (
            "进行中",
            {"fields": ("process_submitted_at", "participants", "handle_method", "expected_plan")},
        ),
        (
            "调解结果",
            {
                "fields": (
                    "result",
                    "result_detail",
                    "process_description",
                    "completed_at",
                    ("complete_lng", "complete_lat"),
                    "complete_address",
                    "complete_image_ids",
                    "complete_file_ids",
                )
            },
        ),
        ("时间", {"fields": ("created_at", "updated_at")}),
    )

    def get_queryset(self, request):
        """排除已归档的任务。"""
        qs = super().get_queryset(request)
        return qs.exclude(status=Task.Status.ARCHIVED)

    @admin.action(description="归档选中的已完成任务")
    def archive_tasks(self, request, queryset):
        """将已完成的任务归档。"""
        # 只归档状态为已完成的任务
        completed_tasks = queryset.filter(status=Task.Status.COMPLETED)
        count = completed_tasks.update(status=Task.Status.ARCHIVED)
        
        if count:
            self.message_user(request, f"成功归档 {count} 条任务")
        else:
            self.message_user(request, "没有可归档的任务（只有已完成状态的任务可以归档）", level="warning")

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "grid":
            kwargs["queryset"] = Grid.objects.filter(is_active=True)
        if db_field.name in {"reporter", "assigned_mediator"}:
            kwargs["queryset"] = User.objects.filter(role=User.Role.MEDIATOR, is_active=True)
        if db_field.name == "assigner":
            kwargs["queryset"] = User.objects.filter(
                role__in=[User.Role.ADMIN, User.Role.GRID_MANAGER],
                is_active=True,
            )
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def save_model(self, request, obj: Task, form, change):
        """
        管理端便捷逻辑：
        - 新增任务时自动生成 code（并发冲突简单重试）
        - 选择调解员后自动写入 assigner/assigned_at，并在已上报状态下切到已分配
        """

        if obj.assigned_mediator_id and (not obj.assigner_id):
            obj.assigner = request.user
        if obj.assigned_mediator_id and (not obj.assigned_at):
            obj.assigned_at = timezone.now()
        if obj.assigned_mediator_id and obj.status == Task.Status.REPORTED:
            obj.status = Task.Status.ASSIGNED
        if obj.process_submitted_at and obj.status == Task.Status.ASSIGNED:
            obj.status = Task.Status.PROCESSING
        if obj.completed_at and obj.status != Task.Status.COMPLETED:
            obj.status = Task.Status.COMPLETED

        if obj.code:
            return super().save_model(request, obj, form, change)

        last_error: IntegrityError | None = None
        for _ in range(3):
            task_type_code = obj.task_type.code if obj.task_type else "task"
            obj.code = generate_task_code(task_type=task_type_code)
            try:
                with transaction.atomic():
                    return super().save_model(request, obj, form, change)
            except IntegrityError as e:
                last_error = e
                obj.code = ""
                continue

        if last_error:
            raise last_error

    def view_detail_action(self, obj):
        """查看详情按钮。"""
        from django.utils.html import format_html
        return format_html(
            '<a style="display:inline-block;padding:4px 12px;background:#e3f2fd;color:#333;'
            'border-radius:4px;text-decoration:none;font-size:12px;" '
            'href="{}">详情</a>',
            f"/admin/cases/task/{obj.pk}/detail/"
        )
    view_detail_action.short_description = "操作"
    view_detail_action.allow_tags = True

    def get_urls(self):
        """添加自定义详情页 URL。"""
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path(
                "<int:task_id>/detail/",
                self.admin_site.admin_view(self.detail_view),
                name="cases_task_detail",
            ),
        ]
        return custom_urls + urls

    def detail_view(self, request, task_id):
        """任务详情视图（管理员端）。"""
        from django.shortcuts import get_object_or_404, render

        task = get_object_or_404(
            Task.objects.select_related("grid", "reporter", "assigned_mediator", "assigner"),
            pk=task_id
        )

        # 获取附件
        report_images = get_attachments_from_ids(task.report_image_ids)
        report_files = get_attachments_from_ids(task.report_file_ids)
        complete_images = get_attachments_from_ids(task.complete_image_ids)
        complete_files = get_attachments_from_ids(task.complete_file_ids)

        # 根据来源确定返回链接
        from_page = request.GET.get("from", "")
        if from_page == "archived":
            back_url = "/admin/cases/archivedtask/"
            back_text = "返回任务归档"
        else:
            back_url = "/admin/cases/task/"
            back_text = "返回任务列表"

        context = {
            "title": f"任务详情: {task.code}",
            "task": task,
            "opts": self.model._meta,
            "has_view_permission": True,
            "back_url": back_url,
            "back_text": back_text,
            "report_images": report_images,
            "report_files": report_files,
            "complete_images": complete_images,
            "complete_files": complete_files,
        }
        return render(request, "admin/cases/task/detail.html", context)


class TaskTypeAdmin(admin.ModelAdmin):
    """任务类型管理。"""

    list_display = ("id", "name", "code", "is_active", "sort_order", "created_at")
    list_editable = ("is_active", "sort_order")
    search_fields = ("name", "code")
    list_filter = ("is_active",)
    ordering = ("sort_order", "id")


class TownAdmin(admin.ModelAdmin):
    """所属镇管理。"""

    list_display = ("id", "name", "code", "is_active", "sort_order", "created_at")
    list_editable = ("is_active", "sort_order")
    search_fields = ("name", "code")
    list_filter = ("is_active",)
    ordering = ("sort_order", "id")


class ArchivedTaskAdmin(admin.ModelAdmin):
    """已归档任务管理（只读，自定义列表页）。"""

    change_list_template = "admin/cases/archivedtask/change_list.html"

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def changelist_view(self, request, extra_context=None):
        """自定义归档任务列表视图。"""
        from django.core.paginator import Paginator
        from django.db.models import Count, Q

        # 获取所有已归档任务的基础查询集
        base_qs = Task.objects.filter(status=Task.Status.ARCHIVED).select_related(
            "grid", "reporter", "assigned_mediator", "task_type", "town"
        )

        # 统计数据：按任务类型分组
        task_types = TaskType.objects.filter(is_active=True).order_by("sort_order", "id")
        stats = []
        colors = ["purple", "green", "orange", "pink", "indigo", "teal"]
        icons = [
            "M9 12l2 2 4-4m6 2a9 9 0 11-18 0 9 9 0 0118 0z",  # 勾选
            "M12 6.253v13m0-13C10.832 5.477 9.246 5 7.5 5S4.168 5.477 3 6.253v13C4.168 18.477 5.754 18 7.5 18s3.332.477 4.5 1.253m0-13C13.168 5.477 14.754 5 16.5 5c1.747 0 3.332.477 4.5 1.253v13C19.832 18.477 18.247 18 16.5 18c-1.746 0-3.332.477-4.5 1.253",  # 书
            "M17 8h2a2 2 0 012 2v6a2 2 0 01-2 2h-2v4l-4-4H9a1.994 1.994 0 01-1.414-.586m0 0L11 14h4a2 2 0 002-2V6a2 2 0 00-2-2H5a2 2 0 00-2 2v6a2 2 0 002 2h2v4l.586-.586z",  # 聊天
        ]
        for i, tt in enumerate(task_types):
            count = base_qs.filter(task_type=tt).count()
            stats.append({
                "id": tt.id,
                "name": tt.name,
                "count": count,
                "color": colors[i % len(colors)],
                "icon": icons[i % len(icons)],
            })

        total_count = base_qs.count()

        # 获取筛选参数
        current_task_type = request.GET.get("task_type")
        if current_task_type:
            try:
                current_task_type = int(current_task_type)
            except (ValueError, TypeError):
                current_task_type = None

        search_query = request.GET.get("q", "").strip()
        current_grid = request.GET.get("grid")
        if current_grid:
            try:
                current_grid = int(current_grid)
            except (ValueError, TypeError):
                current_grid = None

        # 应用筛选
        qs = base_qs
        if current_task_type:
            qs = qs.filter(task_type_id=current_task_type)
        if search_query:
            qs = qs.filter(
                Q(code__icontains=search_query) |
                Q(party_name__icontains=search_query) |
                Q(party_phone__icontains=search_query)
            )
        if current_grid:
            qs = qs.filter(grid_id=current_grid)

        qs = qs.order_by("-completed_at", "-id")

        # 分页
        paginator = Paginator(qs, 20)
        page_number = request.GET.get("page", 1)
        page_obj = paginator.get_page(page_number)

        # 获取网格列表用于筛选
        grids = Grid.objects.filter(is_active=True).order_by("name")

        from django.shortcuts import render

        context = {
            **self.admin_site.each_context(request),
            "title": "任务归档",
            "stats": stats,
            "total_count": total_count,
            "current_task_type": current_task_type,
            "search_query": search_query,
            "current_grid": current_grid,
            "grids": grids,
            "tasks": page_obj.object_list,
            "page_obj": page_obj,
            "opts": self.model._meta,
        }
        return render(request, self.change_list_template, context)


class ExcelImportExportMixin:
    """案件归档 Excel 导入导出 Mixin，提供模板下载和友好错误提示。"""

    excel_template_file = ""  # 子类需要设置模板文件名

    def get_urls(self):
        """添加模板下载 URL。"""
        urls = super().get_urls()
        custom_urls = [
            path(
                "download-template/",
                self.admin_site.admin_view(self.download_template_view),
                name=f"{self.model._meta.app_label}_{self.model._meta.model_name}_download_template",
            ),
        ]
        return custom_urls + urls

    def download_template_view(self, request):
        """下载导入模板文件。"""
        if not self.excel_template_file:
            messages.error(request, "模板文件未配置")
            return redirect("../")

        template_path = getattr(settings, "IMPORT_TEMPLATES_DIR", None)
        if not template_path:
            messages.error(request, "模板目录未配置")
            return redirect("../")

        file_path = os.path.join(template_path, self.excel_template_file)
        if not os.path.exists(file_path):
            messages.error(request, f"模板文件不存在: {self.excel_template_file}")
            return redirect("../")

        response = FileResponse(
            open(file_path, "rb"),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        from urllib.parse import quote
        encoded_filename = quote(self.excel_template_file)
        response["Content-Disposition"] = f"attachment; filename*=UTF-8''{encoded_filename}"
        return response

    def get_import_formats(self):
        """只支持 Excel 格式。"""
        return [XLSX]

    def generate_log_entries(self, result, request):
        """生成导入日志并显示详细错误信息。"""
        if result.has_errors():
            # 收集所有错误信息
            error_messages = []
            for row_errors in result.row_errors():
                row_number, errors = row_errors
                for error in errors:
                    error_msg = str(error.error)
                    error_messages.append(f"第 {row_number + 1} 行: {error_msg}")

            # 显示错误摘要
            if error_messages:
                error_summary = "<br>".join(error_messages[:10])  # 最多显示10条
                if len(error_messages) > 10:
                    error_summary += f"<br>...还有 {len(error_messages) - 10} 条错误"
                messages.error(
                    request,
                    format_html(
                        "导入失败，共 {} 条错误：<br>{}",
                        len(error_messages),
                        format_html(error_summary),
                    ),
                )
        else:
            # 统计导入结果
            new_count = sum(1 for row in result.rows if row.import_type == RowResult.IMPORT_TYPE_NEW)
            update_count = sum(1 for row in result.rows if row.import_type == RowResult.IMPORT_TYPE_UPDATE)
            skip_count = sum(1 for row in result.rows if row.import_type == RowResult.IMPORT_TYPE_SKIP)

            if new_count > 0 or update_count > 0:
                messages.success(
                    request,
                    f"导入成功！新增 {new_count} 条，更新 {update_count} 条，跳过 {skip_count} 条。",
                )

        return super().generate_log_entries(result, request)


class CaseArchiveAdmin(ImportMixin, ExcelImportExportMixin, admin.ModelAdmin):
    """案件归档管理，支持Excel导入导出。"""

    resource_class = CaseArchiveResource
    excel_template_file = "案件归档导入模板.xlsx"
    change_list_template = "admin/cases/casearchive/change_list.html"

    list_display = (
        "id",
        "serial_number",
        "applicant",
        "respondent",
        "case_reason_short",
        "acceptance_time",
        "handler",
        "closure_time",
        "closure_method_short",
        "case_number",
    )
    search_fields = ("applicant", "respondent", "case_number", "handler", "case_reason")
    list_filter = ("acceptance_time", "closure_time")
    ordering = ("-created_at", "-id")
    list_per_page = 20

    readonly_fields = ("created_at", "updated_at")
    fieldsets = (
        ("基本信息", {"fields": ("serial_number", "case_number", "applicant", "respondent")}),
        ("案件详情", {"fields": ("case_reason", "applicable_procedure")}),
        ("办理信息", {"fields": ("handler", "acceptance_time", "closure_time", "closure_method")}),
        ("时间信息", {"fields": ("created_at", "updated_at")}),
    )

    @admin.display(description="案由")
    def case_reason_short(self, obj):
        """案由截断显示。"""
        if obj.case_reason and len(obj.case_reason) > 20:
            return f"{obj.case_reason[:20]}..."
        return obj.case_reason

    @admin.display(description="结案方式")
    def closure_method_short(self, obj):
        """结案方式截断显示。"""
        if obj.closure_method and len(obj.closure_method) > 15:
            return f"{obj.closure_method[:15]}..."
        return obj.closure_method


class TaskStatReportAdmin(admin.ModelAdmin):
    """统计报表：按月统计任务，按镇办×矛盾纠纷类型交叉统计，支持导出 Excel。"""

    change_list_template = "admin/cases/taskstatreport/change_list.html"

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def _get_month_range(self, year, month):
        """Return (start_date, end_date) for a given month."""
        import calendar
        from datetime import date
        last_day = calendar.monthrange(year, month)[1]
        return date(year, month, 1), date(year, month, last_day)

    def _build_stat_data(self, year, month):
        """构建统计数据: 按镇办×类型交叉统计化解/未化解数。"""
        from django.db.models import Count, Q

        start_date, end_date = self._get_month_range(year, month)

        # 基础查询：该月上报的任务
        base_qs = Task.objects.filter(
            reported_at__date__gte=start_date,
            reported_at__date__lte=end_date,
        )

        towns = list(Town.objects.filter(is_active=True).order_by("sort_order", "id"))
        task_types = list(TaskType.objects.filter(is_active=True).order_by("sort_order", "id"))

        # 化解 = success + partial，未化解 = 其他（包含 failure 及未完成的）
        resolved_q = Q(result__in=[Task.Result.SUCCESS, Task.Result.PARTIAL])

        # 按 town×task_type 统计
        raw = (
            base_qs
            .values("town_id", "task_type_id")
            .annotate(
                total=Count("id"),
                resolved=Count("id", filter=resolved_q),
            )
        )
        # 构建查找表 {(town_id, task_type_id): {total, resolved}}
        lookup = {}
        for row in raw:
            lookup[(row["town_id"], row["task_type_id"])] = {
                "total": row["total"],
                "resolved": row["resolved"],
            }

        # 每个类型的全局统计
        type_totals = {}
        for tt in task_types:
            type_totals[tt.id] = {"resolved": 0, "unresolved": 0}

        # 构建每行数据
        rows = []
        grand_total = 0
        grand_unresolved = 0
        for town in towns:
            row_total = 0
            row_unresolved = 0
            cells = []
            for tt in task_types:
                info = lookup.get((town.id, tt.id), {"total": 0, "resolved": 0})
                resolved = info["resolved"]
                unresolved = info["total"] - resolved
                cells.append({"resolved": resolved, "unresolved": unresolved})
                row_total += info["total"]
                row_unresolved += unresolved
                type_totals[tt.id]["resolved"] += resolved
                type_totals[tt.id]["unresolved"] += unresolved
            rows.append({
                "town": town,
                "total": row_total,
                "unresolved": row_unresolved,
                "cells": cells,
            })
            grand_total += row_total
            grand_unresolved += row_unresolved

        # 合计行 + 类型统计（zip task_type with summary for template rendering）
        type_stats = []
        summary_cells = []
        for tt in task_types:
            info = type_totals[tt.id]
            info["total"] = info["resolved"] + info["unresolved"]
            summary_cells.append(info)
            type_stats.append({"type": tt, **info})

        grand_resolved = grand_total - grand_unresolved

        return {
            "towns": towns,
            "task_types": task_types,
            "type_stats": type_stats,
            "rows": rows,
            "summary_cells": summary_cells,
            "grand_total": grand_total,
            "grand_resolved": grand_resolved,
            "grand_unresolved": grand_unresolved,
            "year": year,
            "month": month,
        }

    def changelist_view(self, request, extra_context=None):
        """Render the monthly statistics report page."""
        from django.shortcuts import render

        now = timezone.now()
        try:
            year = int(request.GET.get("year", now.year))
            month = int(request.GET.get("month", now.month))
            if not (1 <= month <= 12) or year < 2000:
                raise ValueError
        except (ValueError, TypeError):
            year, month = now.year, now.month

        stat = self._build_stat_data(year, month)

        # 构建年/月选项
        year_choices = list(range(2024, now.year + 2))
        month_choices = list(range(1, 13))

        context = {
            **self.admin_site.each_context(request),
            "title": "统计报表",
            "opts": self.model._meta,
            "stat": stat,
            "year_choices": year_choices,
            "month_choices": month_choices,
            "selected_year": year,
            "selected_month": month,
        }
        return render(request, self.change_list_template, context)

    # 模板文件名（位于 IMPORT_TEMPLATES_DIR 目录下）
    excel_template_file = "矛盾纠纷统计报表模板.xlsx"

    @staticmethod
    def _copy_style(src, dst):
        """Copy font/fill/border/alignment/number_format from src cell to dst cell."""
        from copy import copy
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format

    @staticmethod
    def _ensure_borders(ws, max_row, max_col, start_row=1):
        """确保数据区域所有单元格都有完整边框（保留已有边框样式，缺失的补 thin）。"""
        from copy import copy
        from openpyxl.styles import Border, Side
        thin = Side(style="thin")
        for r in range(start_row, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(row=r, column=c)
                b = cell.border
                left = b.left if b.left and b.left.style else thin
                right = b.right if b.right and b.right.style else thin
                top = b.top if b.top and b.top.style else thin
                bottom = b.bottom if b.bottom and b.bottom.style else thin
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

    def _export_excel(self, request, year, month):
        """
        加载模板文件，读取样式参考单元格（D-E 列），填充实际数据后导出。

        模板行号约定（含标题行）：
          Row 1 — 标题行       Row 2 — 镇办/矛盾纠纷类型
          Row 3 — D3: 类型表头  Row 4 — D4/E4: 子表头
          Row 5 — D5/E5: 合计行  Row 6 — D6/E6: 数据行
        可用 Excel 编辑模板文件自由调整样式。
        """
        from io import BytesIO
        from urllib.parse import quote

        from django.http import HttpResponse
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        stat = self._build_stat_data(year, month)
        task_types = stat["task_types"]
        rows_data = stat["rows"]
        n_types = len(task_types)

        # --- 加载模板 ---
        tpl_path = os.path.join(settings.IMPORT_TEMPLATES_DIR, self.excel_template_file)
        wb = load_workbook(tpl_path)
        ws = wb.active
        ws.title = f"{year}年{month}月统计"

        # --- 从模板读取参考样式（模板含标题行，行号 1-6） ---
        ref_title       = ws.cell(row=1, column=1)   # A1: 标题行
        ref_type_header = ws.cell(row=3, column=4)   # D3: 类型表头
        ref_sub_left    = ws.cell(row=4, column=4)   # D4: 化解数表头
        ref_sub_right   = ws.cell(row=4, column=5)   # E4: 未化解数表头
        ref_sum_left    = ws.cell(row=5, column=4)   # D5: 合计行左
        ref_sum_right   = ws.cell(row=5, column=5)   # E5: 合计行右
        ref_data_left   = ws.cell(row=6, column=4)   # D6: 数据行左
        ref_data_right  = ws.cell(row=6, column=5)   # E6: 数据行右
        # 固定列参考
        ref_a5  = ws.cell(row=5, column=1)  # 合计行镇办列
        ref_b5  = ws.cell(row=5, column=2)  # 合计行总数列
        ref_c5  = ws.cell(row=5, column=3)
        ref_a6  = ws.cell(row=6, column=1)  # 数据行镇办列
        ref_b6  = ws.cell(row=6, column=2)
        ref_c6  = ws.cell(row=6, column=3)

        # 行高参考
        row_heights = {}
        for r in range(1, 5):
            h = ws.row_dimensions[r].height
            if h:
                row_heights[r] = h

        # --- 清空模板数据区域 ---
        for mg in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(mg))
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.value = None

        total_cols = 3 + n_types * 2

        # --- Row 1: 标题行 ---
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        c = ws.cell(row=1, column=1, value=f"镇巴县司法行政系统矛盾纠纷排查化解基本情况统计表（{month} 月）")
        self._copy_style(ref_title, c)

        # --- Row 2-4: 表头 ---
        ws.merge_cells(start_row=2, start_column=1, end_row=4, end_column=1)
        ws.cell(row=2, column=1, value="镇办")

        if total_cols > 1:
            ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=total_cols)
            ws.cell(row=2, column=2, value="矛盾纠纷类型")

        # Row 3
        ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=3)
        ws.cell(row=3, column=2, value="当前矛盾\n纠纷情况")

        for i, tt in enumerate(task_types):
            col = 4 + i * 2
            tt_total = stat["summary_cells"][i]["total"]
            ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 1)
            c = ws.cell(row=3, column=col, value=f"{tt.name}\n（共计：{tt_total} 件）")
            self._copy_style(ref_type_header, c)

        # Row 4: 子表头
        ws.cell(row=4, column=2, value="本月矛盾\n纠纷总数")
        ws.cell(row=4, column=3, value="本月未化解\n纠纷数")
        for i in range(n_types):
            col = 4 + i * 2
            c = ws.cell(row=4, column=col, value="化解数")
            self._copy_style(ref_sub_left, c)
            c = ws.cell(row=4, column=col + 1, value="未化解数")
            self._copy_style(ref_sub_right, c)

        # --- Row 5: 合计 ---
        c = ws.cell(row=5, column=1, value="合计")
        self._copy_style(ref_a5, c)
        c = ws.cell(row=5, column=2, value=stat["grand_total"])
        self._copy_style(ref_b5, c)
        c = ws.cell(row=5, column=3, value=stat["grand_unresolved"])
        self._copy_style(ref_c5, c)
        for i, sc in enumerate(stat["summary_cells"]):
            col = 4 + i * 2
            c = ws.cell(row=5, column=col, value=sc["resolved"])
            self._copy_style(ref_sum_left, c)
            c = ws.cell(row=5, column=col + 1, value=sc["unresolved"])
            self._copy_style(ref_sum_right, c)

        # --- Data rows (from row 6) ---
        for idx, row_data in enumerate(rows_data):
            r = 6 + idx
            c = ws.cell(row=r, column=1, value=row_data["town"].name)
            self._copy_style(ref_a6, c)
            c = ws.cell(row=r, column=2, value=row_data["total"])
            self._copy_style(ref_b6, c)
            c = ws.cell(row=r, column=3, value=row_data["unresolved"])
            self._copy_style(ref_c6, c)
            for i, cell_data in enumerate(row_data["cells"]):
                col = 4 + i * 2
                c = ws.cell(row=r, column=col, value=cell_data["resolved"])
                self._copy_style(ref_data_left, c)
                c = ws.cell(row=r, column=col + 1, value=cell_data["unresolved"])
                self._copy_style(ref_data_right, c)

        # --- 列宽、行高 ---
        ws.column_dimensions["A"].width = 14
        for col_idx in range(2, total_cols + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 13
        for row_num, h in row_heights.items():
            ws.row_dimensions[row_num].height = h

        # --- 补全边框（标题行不加边框，从第 2 行开始） ---
        max_data_row = 5 + len(rows_data)
        self._ensure_borders(ws, max_data_row, total_cols, start_row=2)

        # --- 输出 ---
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"{year}年{month}月矛盾纠纷统计报表.xlsx"
        response = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quote(filename)}"
        return response

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                "export/",
                self.admin_site.admin_view(self.export_view),
                name="cases_taskstatreport_export",
            ),
        ]
        return custom_urls + urls

    def export_view(self, request):
        """Handle Excel export request."""
        now = timezone.now()
        try:
            year = int(request.GET.get("year", now.year))
            month = int(request.GET.get("month", now.month))
            if not (1 <= month <= 12) or year < 2000:
                raise ValueError
        except (ValueError, TypeError):
            year, month = now.year, now.month
        return self._export_excel(request, year, month)


# 注册到管理员后台
admin_site.register(TaskType, TaskTypeAdmin)
admin_site.register(Town, TownAdmin)
admin_site.register(Task, TaskAdmin)
admin_site.register(ArchivedTask, ArchivedTaskAdmin)
admin_site.register(CaseArchive, CaseArchiveAdmin)
admin_site.register(TaskStatReport, TaskStatReportAdmin)

# ==================== 网格管理员专用 Admin ====================

class GridManagerTaskAdmin(admin.ModelAdmin):
    """
    网格管理员端 - 所有任务列表。

    功能：
    - 查看本网格所有任务
    - 点击查看任务详情（自定义页面）
    """

    list_display = (
        "id",
        "code",
        "task_type",
        "status",
        "party_name",
        "reporter",
        "assigned_mediator",
        "reported_at",
        "view_detail_action",
    )
    list_display_links = None  # 禁用ID点击进入详情
    list_select_related = ("grid", "reporter", "assigned_mediator", "task_type")
    search_fields = ("code", "party_name", "party_phone", "reporter__name", "assigned_mediator__name")
    list_filter = ("task_type", "status", "reported_at")
    date_hierarchy = "reported_at"
    ordering = ("-reported_at", "-id")

    def get_queryset(self, request):
        """只显示本网格的任务。"""
        queryset = super().get_queryset(request)
        managed_grids = Grid.objects.filter(current_manager=request.user, is_active=True)
        return queryset.filter(grid__in=managed_grids)

    def view_detail_action(self, obj):
        """查看详情按钮。"""
        from django.utils.html import format_html
        return format_html(
            '<a style="display:inline-block;padding:4px 12px;background:#e3f2fd;color:#333;'
            'border-radius:4px;text-decoration:none;font-size:12px;" '
            'href="{}">详情</a>',
            f"/grid-admin/cases/task/{obj.pk}/detail/"
        )
    view_detail_action.short_description = "操作"
    view_detail_action.allow_tags = True

    def has_add_permission(self, request):
        """网格管理员不能新增任务（任务由调解员上报）。"""
        return False

    def has_change_permission(self, request, obj=None):
        """网格管理员不能编辑任务（只能查看）。"""
        return False

    def has_delete_permission(self, request, obj=None):
        """网格管理员不能删除任务。"""
        return False

    def get_urls(self):
        """添加自定义详情页 URL。"""
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path(
                "<int:task_id>/detail/",
                self.admin_site.admin_view(self.detail_view),
                name="task_detail",
            ),
        ]
        return custom_urls + urls

    def detail_view(self, request, task_id):
        """任务详情视图。"""
        from django.shortcuts import get_object_or_404, render
        from django.contrib import messages

        task = get_object_or_404(
            Task.objects.select_related("grid", "reporter", "assigned_mediator", "assigner"),
            pk=task_id
        )

        # 检查权限：任务必须属于当前管理员管理的网格
        managed_grids = Grid.objects.filter(current_manager=request.user, is_active=True)
        if task.grid not in managed_grids:
            messages.error(request, "您没有权限查看此任务")
            from django.shortcuts import redirect
            return redirect("grid_admin:cases_task_changelist")

        # 根据来源确定返回链接
        from_page = request.GET.get("from", "")
        if from_page == "unassigned":
            back_url = "/grid-admin/cases/unassignedtask/"
            back_text = "返回待分配任务"
        else:
            back_url = "/grid-admin/cases/task/"
            back_text = "返回任务列表"

        # 获取附件
        report_images = get_attachments_from_ids(task.report_image_ids)
        report_files = get_attachments_from_ids(task.report_file_ids)
        complete_images = get_attachments_from_ids(task.complete_image_ids)
        complete_files = get_attachments_from_ids(task.complete_file_ids)

        context = {
            "title": f"任务详情: {task.code}",
            "task": task,
            "opts": self.model._meta,
            "has_view_permission": True,
            "back_url": back_url,
            "back_text": back_text,
            "report_images": report_images,
            "report_files": report_files,
            "complete_images": complete_images,
            "complete_files": complete_files,
        }
        return render(request, "admin/cases/task/detail.html", context)


class AssignMediatorForm(forms.Form):
    """分配调解员表单。"""
    mediator = forms.ModelChoiceField(
        queryset=User.objects.none(),
        label="选择调解员",
        widget=forms.Select(attrs={"class": "form-control"}),
    )

    def __init__(self, *args, grid=None, **kwargs):
        super().__init__(*args, **kwargs)
        if grid:
            self.fields["mediator"].queryset = User.objects.filter(
                role=User.Role.MEDIATOR, is_active=True, grid=grid
            )


class GridManagerUnassignedTaskAdmin(admin.ModelAdmin):
    """
    网格管理员端 - 未分配任务列表。

    功能：
    - 查看本网格未分配的任务（状态为 reported）
    - 点击分配按钮分配给本网格调解员
    """

    list_display = (
        "id",
        "code",
        "task_type",
        "party_name",
        "reporter",
        "reported_at",
        "action_buttons",
    )
    list_display_links = None  # 禁用ID点击进入详情
    list_select_related = ("grid", "reporter", "task_type")
    search_fields = ("code", "party_name", "party_phone", "reporter__name")
    list_filter = ("task_type", "reported_at")
    date_hierarchy = "reported_at"
    ordering = ("-reported_at", "-id")

    def get_queryset(self, request):
        """只显示本网格未分配的任务。"""
        queryset = super().get_queryset(request)
        managed_grids = Grid.objects.filter(current_manager=request.user, is_active=True)
        return queryset.filter(grid__in=managed_grids, status=Task.Status.REPORTED)

    def action_buttons(self, obj):
        """操作按钮：详情 + 分配。"""
        from django.utils.html import format_html
        btn_style = (
            'display:inline-block;padding:4px 12px;background:#e3f2fd;color:#333;'
            'border-radius:4px;text-decoration:none;font-size:12px;margin-right:8px;'
        )
        return format_html(
            '<a style="{}" href="{}">详情</a>'
            '<a style="{}" href="{}">分配</a>',
            btn_style,
            f"/grid-admin/cases/task/{obj.pk}/detail/?from=unassigned",
            btn_style,
            f"/grid-admin/cases/unassignedtask/{obj.pk}/assign/"
        )
    action_buttons.short_description = "操作"
    action_buttons.allow_tags = True

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def get_urls(self):
        """添加自定义分配页面 URL。"""
        from django.urls import path
        urls = super().get_urls()
        custom_urls = [
            path(
                "<int:task_id>/assign/",
                self.admin_site.admin_view(self.assign_view),
                name="unassignedtask_assign",
            ),
        ]
        return custom_urls + urls

    def assign_view(self, request, task_id):
        """分配任务视图。"""
        from django.shortcuts import get_object_or_404, redirect, render
        from django.contrib import messages

        task = get_object_or_404(Task, pk=task_id)

        # 检查权限：任务必须属于当前管理员管理的网格
        managed_grids = Grid.objects.filter(current_manager=request.user, is_active=True)
        if task.grid not in managed_grids:
            messages.error(request, "您没有权限分配此任务")
            return redirect("grid_admin:cases_unassignedtask_changelist")

        if task.status != Task.Status.REPORTED:
            messages.error(request, "此任务已被分配")
            return redirect("grid_admin:cases_unassignedtask_changelist")

        if request.method == "POST":
            form = AssignMediatorForm(request.POST, grid=task.grid)
            if form.is_valid():
                mediator = form.cleaned_data["mediator"]
                task.assigned_mediator = mediator
                task.assigner = request.user
                task.assigned_at = timezone.now()
                task.status = Task.Status.ASSIGNED
                task.save(update_fields=["assigned_mediator", "assigner", "assigned_at", "status", "updated_at"])
                messages.success(request, f"任务 {task.code} 已分配给 {mediator.name}")
                return redirect("grid_admin:cases_unassignedtask_changelist")
        else:
            form = AssignMediatorForm(grid=task.grid)

        # 获取上报附件
        report_images = get_attachments_from_ids(task.report_image_ids)
        report_files = get_attachments_from_ids(task.report_file_ids)

        context = {
            "title": f"分配任务: {task.code}",
            "task": task,
            "form": form,
            "opts": self.model._meta,
            "has_view_permission": True,
            "report_images": report_images,
            "report_files": report_files,
        }
        return render(request, "admin/cases/unassignedtask/assign.html", context)


# 注册到网格负责人后台
grid_manager_site.register(Task, GridManagerTaskAdmin)
grid_manager_site.register(UnassignedTask, GridManagerUnassignedTaskAdmin)
